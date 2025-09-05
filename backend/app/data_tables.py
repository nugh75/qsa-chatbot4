"""
Data Tables module: upload CSV/XLSX into managed tables, CRUD rows, and search.

Storage model:
- Metadata in table `data_tables`
- Rows stored as JSON per-row in `data_table_rows`
- Original uploaded file stored under storage/data_tables/originals/{table_id}/

This module works with both SQLite and PostgreSQL via the shared db_manager.
It creates the necessary tables if missing (idempotent) for both engines.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from fastapi import HTTPException

from .database import db_manager, USING_POSTGRES


BASE_DIR = Path(__file__).resolve().parent.parent
TABLES_BASE_DIR = BASE_DIR / "storage" / "data_tables"
ORIGINALS_DIR = TABLES_BASE_DIR / "originals"

for d in (TABLES_BASE_DIR, ORIGINALS_DIR):
    try:
        d.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


def _slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9\-\s]", "", s)
    s = re.sub(r"\s+", "-", s)
    return s or f"tbl-{uuid.uuid4().hex[:8]}"


def init_tables_schema():
    """Create data tables schema if missing (SQLite and Postgres)."""
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        if USING_POSTGRES:
            # Metadata
            db_manager.exec(cur, """
                CREATE TABLE IF NOT EXISTS data_tables (
                  id TEXT PRIMARY KEY,
                  name TEXT UNIQUE NOT NULL,
                  title TEXT NOT NULL,
                  description TEXT,
                  columns JSONB,
                  original_filename TEXT,
                  file_format TEXT,
                  row_count INTEGER NOT NULL DEFAULT 0,
                  created_by INTEGER NULL,
                  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                  updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                  CONSTRAINT fk_dt_user FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
                )
            """)
            # Rows
            db_manager.exec(cur, """
                CREATE TABLE IF NOT EXISTS data_table_rows (
                  id TEXT PRIMARY KEY,
                  table_id TEXT NOT NULL REFERENCES data_tables(id) ON DELETE CASCADE,
                  data JSONB NOT NULL,
                  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                  updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            db_manager.exec(cur, "CREATE INDEX IF NOT EXISTS idx_dtr_table ON data_table_rows(table_id)")
        else:
            # SQLite
            cur.execute("""
                CREATE TABLE IF NOT EXISTS data_tables (
                    id TEXT PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    title TEXT NOT NULL,
                    description TEXT,
                    columns TEXT,
                    original_filename TEXT,
                    file_format TEXT,
                    row_count INTEGER DEFAULT 0,
                    created_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS data_table_rows (
                    id TEXT PRIMARY KEY,
                    table_id TEXT NOT NULL,
                    data TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (table_id) REFERENCES data_tables(id) ON DELETE CASCADE
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_dtr_table ON data_table_rows(table_id)")
        conn.commit()


def _now_str() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _detect_columns_from_rows(rows: Iterable[Dict[str, Any]], max_rows: int = 1000) -> List[str]:
    cols: List[str] = []
    seen = set()
    for i, r in enumerate(rows):
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
        if i >= max_rows:
            break
    return cols


def _read_csv(file_bytes: bytes, encoding: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    data = file_bytes.decode(encoding or 'utf-8', errors='replace')
    f = io.StringIO(data)
    # Try to sniff delimiter
    try:
        dialect = csv.Sniffer().sniff(data[:2048])
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(f, dialect=dialect)
    rows = [dict(row) for row in reader]
    cols = reader.fieldnames or _detect_columns_from_rows(rows)
    return (cols or []), rows


def _read_xlsx(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Supporto Excel non disponibile: installare openpyxl ({e})")
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []
    cols = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(header)]
    rows: List[Dict[str, Any]] = []
    for r in rows_iter:
        obj: Dict[str, Any] = {}
        for i, c in enumerate(cols):
            val = r[i] if i < len(r) else None
            if isinstance(val, datetime):
                val = val.isoformat()
            obj[c] = val
        if any(v is not None and str(v) != '' for v in obj.values()):
            rows.append(obj)
    return cols, rows


def _write_csv(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    s = io.StringIO()
    writer = csv.DictWriter(s, fieldnames=columns)
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k, '') for k in columns})
    return s.getvalue().encode('utf-8')


def _write_xlsx(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Supporto Excel non disponibile: installare openpyxl ({e})")
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, '') for c in columns])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def create_table_from_upload(
    *,
    title: str,
    description: Optional[str],
    filename: str,
    file_bytes: bytes,
    created_by_user_id: Optional[int] = None,
    preferred_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Parse CSV/XLSX and create a new managed table with rows."""
    init_tables_schema()
    name = _slugify(preferred_name or title or Path(filename).stem)
    ext = Path(filename).suffix.lower().lstrip('.')
    if ext in ("csv", "tsv"):
        cols, rows = _read_csv(file_bytes)
        fmt = "csv"
    elif ext in ("xlsx", "xlsm", "xltx", "xltm", "xls"):
        cols, rows = _read_xlsx(file_bytes)
        fmt = "xlsx"
    else:
        # Try CSV fallback
        try:
            cols, rows = _read_csv(file_bytes)
            fmt = "csv"
        except Exception:
            raise HTTPException(status_code=400, detail=f"Formato file non supportato: {ext}")

    table_id = f"tbl_{uuid.uuid4().hex}"
    columns_json = json.dumps(cols or [])
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        if USING_POSTGRES:
            db_manager.exec(cur, """
                INSERT INTO data_tables (id, name, title, description, columns, original_filename, file_format, row_count, created_by)
                VALUES (?, ?, ?, ?, %s::jsonb, ?, ?, ?, ?)
            """, (table_id, name, title, description or None, columns_json, filename, fmt, len(rows), created_by_user_id))
        else:
            cur.execute(
                """
                INSERT INTO data_tables (id, name, title, description, columns, original_filename, file_format, row_count, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (table_id, name, title, description or None, columns_json, filename, fmt, len(rows), created_by_user_id),
            )
        # Insert rows
        for r in rows:
            rid = f"row_{uuid.uuid4().hex}"
            if USING_POSTGRES:
                db_manager.exec(cur, "INSERT INTO data_table_rows (id, table_id, data) VALUES (?, ?, %s::jsonb)", (rid, table_id, json.dumps(r, ensure_ascii=False)))
            else:
                cur.execute("INSERT INTO data_table_rows (id, table_id, data) VALUES (?, ?, ?)", (rid, table_id, json.dumps(r, ensure_ascii=False)))
        conn.commit()

    # Save original file for reference
    try:
        td = ORIGINALS_DIR / table_id
        td.mkdir(parents=True, exist_ok=True)
        with open(td / filename, 'wb') as f:
            f.write(file_bytes)
    except Exception:
        pass

    return {
        "id": table_id,
        "name": name,
        "title": title,
        "description": description or "",
        "columns": cols,
        "row_count": len(rows),
        "file_format": fmt,
        "original_filename": filename,
        "created_by": created_by_user_id,
    }


def list_tables() -> List[Dict[str, Any]]:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        db_manager.exec(cur, "SELECT id, name, title, description, original_filename, file_format, row_count, created_at, updated_at FROM data_tables ORDER BY title")
        rows = cur.fetchall()
        out = []
        for r in rows:
            d = dict(r)
            out.append(d)
        return out


def get_table(table_id: str) -> Optional[Dict[str, Any]]:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        db_manager.exec(cur, "SELECT * FROM data_tables WHERE id = ?", (table_id,))
        row = cur.fetchone()
        if not row:
            return None
        d = dict(row)
        cols = d.get('columns')
        if isinstance(cols, (bytes, str)):
            try:
                d['columns'] = json.loads(cols)
            except Exception:
                d['columns'] = []
        return d


def get_rows(table_id: str, limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        db_manager.exec(cur, "SELECT id, data, created_at, updated_at FROM data_table_rows WHERE table_id = ? ORDER BY created_at ASC LIMIT ? OFFSET ?", (table_id, limit, offset))
        rows = cur.fetchall()
        out: List[Dict[str, Any]] = []
        for r in rows:
            d = dict(r)
            payload = d.get('data')
            if isinstance(payload, (bytes, str)):
                try:
                    d['data'] = json.loads(payload)
                except Exception:
                    d['data'] = {}
            out.append(d)
        return out


def delete_table(table_id: str) -> bool:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        db_manager.exec(cur, "DELETE FROM data_table_rows WHERE table_id = ?", (table_id,))
        db_manager.exec(cur, "DELETE FROM data_tables WHERE id = ?", (table_id,))
        conn.commit()
    # Remove originals dir
    try:
        td = ORIGINALS_DIR / table_id
        if td.exists():
            for p in td.glob('*'):
                try:
                    p.unlink()
                except Exception:
                    pass
            try:
                td.rmdir()
            except Exception:
                pass
    except Exception:
        pass
    return True


def update_table_meta(table_id: str, *, title: Optional[str] = None, description: Optional[str] = None) -> None:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        sets = []
        params: List[Any] = []
        if title is not None:
            sets.append("title = ?")
            params.append(title)
        if description is not None:
            sets.append("description = ?")
            params.append(description)
        if not sets:
            return
        if USING_POSTGRES:
            sets.append("updated_at = NOW()")
        else:
            sets.append("updated_at = CURRENT_TIMESTAMP")
        sql = "UPDATE data_tables SET " + ", ".join(sets) + " WHERE id = ?"
        params.append(table_id)
        db_manager.exec(cur, sql, tuple(params))
        conn.commit()


def add_rows(table_id: str, rows: List[Dict[str, Any]]) -> int:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        count = 0
        for r in rows:
            rid = f"row_{uuid.uuid4().hex}"
            payload = json.dumps(r, ensure_ascii=False)
            if USING_POSTGRES:
                db_manager.exec(cur, "INSERT INTO data_table_rows (id, table_id, data) VALUES (?, ?, %s::jsonb)", (rid, table_id, payload))
            else:
                cur.execute("INSERT INTO data_table_rows (id, table_id, data) VALUES (?, ?, ?)", (rid, table_id, payload))
            count += 1
        # Update row_count
        db_manager.exec(cur, "UPDATE data_tables SET row_count = row_count + ? WHERE id = ?", (count, table_id))
        conn.commit()
        return count


def update_row(table_id: str, row_id: str, data: Dict[str, Any]) -> None:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        payload = json.dumps(data, ensure_ascii=False)
        if USING_POSTGRES:
            db_manager.exec(cur, "UPDATE data_table_rows SET data = %s::jsonb, updated_at = NOW() WHERE id = ? AND table_id = ?", (payload, row_id, table_id))
        else:
            cur.execute("UPDATE data_table_rows SET data = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ? AND table_id = ?", (payload, row_id, table_id))
        conn.commit()


def delete_row(table_id: str, row_id: str) -> None:
    init_tables_schema()
    with db_manager.get_connection() as conn:
        cur = conn.cursor()
        db_manager.exec(cur, "DELETE FROM data_table_rows WHERE id = ? AND table_id = ?", (row_id, table_id))
        # Decrement row_count (clamped at 0)
        db_manager.exec(cur, "UPDATE data_tables SET row_count = CASE WHEN row_count > 0 THEN row_count - 1 ELSE 0 END WHERE id = ?", (table_id,))
        conn.commit()


def export_table(table_id: str, fmt: str = "csv", limit: Optional[int] = None) -> Tuple[str, bytes]:
    """Export current table rows as CSV/XLSX.

    Returns a tuple (mime_type, data_bytes)
    """
    t = get_table(table_id)
    if not t:
        raise HTTPException(status_code=404, detail="Tabella non trovata")
    columns: List[str] = t.get('columns') or []
    rows = get_rows(table_id, limit=limit or 1000000, offset=0)
    payload_rows = [r.get('data') or {} for r in rows]
    if fmt == 'xlsx':
        data = _write_xlsx(columns, payload_rows)
        return ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", data)
    else:
        data = _write_csv(columns, payload_rows)
        return ("text/csv; charset=utf-8", data)


def search_tables(query: str, table_ids: List[str], limit_per_table: int = 10) -> Dict[str, Any]:
    """Search across JSON rows for selected tables with light Italian-friendly normalization.

    - Tokenizes query and expands simple variants (singolare/plurale, prefisso >= 5 chars)
    - Recognizes Italian months (e.g., 'settembre' -> {'09','9','settembre','sett'})
    - Ranks by count of distinct matched tokens/variants per row
    """
    # Tokenize
    raw = (query or '').lower()
    tokens = [t for t in re.findall(r"[\w]+", raw) if len(t) > 2]
    if not tokens:
        return {"results": []}

    # Expand token variants
    MONTHS = {
        'gennaio': ['01', '1', 'gen', 'genn'],
        'febbraio': ['02', '2', 'feb'],
        'marzo': ['03', '3', 'mar'],
        'aprile': ['04', '4', 'apr'],
        'maggio': ['05', '5', 'mag'],
        'giugno': ['06', '6', 'giu'],
        'luglio': ['07', '7', 'lug'],
        'agosto': ['08', '8', 'ago'],
        'settembre': ['09', '9', 'set', 'sett'],
        'ottobre': ['10', 'ott'],
        'novembre': ['11', 'nov'],
        'dicembre': ['12', 'dic'],
    }
    def expand(tok: str) -> List[str]:
        variants = {tok}
        # Months
        if tok in MONTHS:
            variants.update(MONTHS[tok])
        # Simple plural/singular heuristics
        if tok.endswith('i'):
            variants.add(tok[:-1])  # lezioni -> lezion
            variants.add(tok[:-1] + 'o')  # lezioni -> lezione (approx)
            variants.add(tok[:-1] + 'e')  # libri -> libre (approx fallback)
        if tok.endswith('e'):
            variants.add(tok[:-1])  # lezione -> lezion
        # Prefix (>=5 chars) for fuzzy match
        if len(tok) >= 5:
            variants.add(tok[:5])
        return [v for v in variants if v and len(v) >= 3]

    variant_sets: List[set] = [set(expand(t)) for t in sorted(set(tokens))]

    results: List[Dict[str, Any]] = []
    for tid in table_ids:
        # Scan all rows of the table to ensure the agent can consider the full dataset
        rows = get_rows(tid, limit=1_000_000, offset=0)
        scored: List[Tuple[int, Dict[str, Any]]] = []
        for r in rows:
            payload = r.get('data') or {}
            s = json.dumps(payload, ensure_ascii=False).lower()
            # Count distinct token matches (any variant per token)
            score = 0
            for vset in variant_sets:
                if any(v in s for v in vset):
                    score += 1
            if score > 0:
                scored.append((score, r))
        scored.sort(key=lambda x: x[0], reverse=True)
        top = [item[1] for item in scored[:limit_per_table]]
        if top:
            tmeta = get_table(tid) or {"id": tid, "title": tid, "name": tid}
            # compute display_columns: first 5 columns
            display_columns = (tmeta.get('columns') or [])[:5]
            results.append({
                "table_id": tid,
                "table_name": tmeta.get('name'),
                "title": tmeta.get('title'),
                "columns": tmeta.get('columns') or [],
                "display_columns": display_columns,
                "rows": top,
            })
    return {"results": results}
